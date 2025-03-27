# extract_transform.py
import os
import logging
import abc
from typing import List, Dict, Any, Optional, Type, Union, Tuple, Callable
from pathlib import Path
import PyPDF2 # Requires PyPDF2 >= 3.0 for get_object robustness, tested up to ~3.x
from openpyxl import load_workbook
try:
    import geopandas
    GEOPANDAS_AVAILABLE = True
except ImportError:
    GEOPANDAS_AVAILABLE = False
    # Define dummy placeholders if geopandas is not available
    class GeoDataFrame: pass
    class geopandas: read_file = lambda *args, **kwargs: GeoDataFrame() # type: ignore
import chardet
import zipfile
import shutil
import datetime
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed # Only for internal zip processing

# --- Dependency Import ---
try:
    from utils import calculate_file_hash
except ImportError:
    logging.warning("utils.calculate_file_hash not found. Using dummy hash implementation.")
    # Define a simple dummy hash function if utils is missing
    calculate_file_hash = lambda filepath, **kwargs: f"dummy_hash_for_{Path(filepath).name}" # type: ignore

log = logging.getLogger(__name__)

# --- Constants ---
DEFAULT_TEMP_DIR = Path("./temp_extraction").resolve()
DEFAULT_MAX_WORKERS_ZIP = os.cpu_count() or 2 # Workers for files *within* a zip

# --- File Handler Mapping ---
SUPPORTED_FILE_EXTENSIONS: Dict[str, Type['FileSystemItem']] = {} # Populated after class defs

# --- Helper Functions ---
def _safe_isoformat(dt: Optional[Any]) -> Optional[str]:
    """Safely convert datetime object to ISO UTC format string."""
    if isinstance(dt, datetime.datetime):
        if dt.tzinfo is None: dt = dt.replace(tzinfo=datetime.timezone.utc) # Assume UTC if naive
        return dt.isoformat()
    return None

def _parse_pdf_date_simple(date_str: Optional[Any]) -> Optional[str]:
    """Basic PDF date string parser -> ISO format (assumes UTC)."""
    if not isinstance(date_str, str) or not date_str.startswith("D:") or len(date_str) < 9:
        return str(date_str) if date_str is not None else None
    pdf_date_str = date_str[2:]
    core_dt_str = pdf_date_str[:14] # YYYYMMDDHHMMSS
    for fmt in ('%Y%m%d%H%M%S', '%Y%m%d%H%M', '%Y%m%d%H', '%Y%m%d'):
        if len(core_dt_str) >= len(fmt.replace('%', '')):
            try:
                parsed_dt = datetime.datetime.strptime(core_dt_str[:len(fmt.replace('%', ''))], fmt)
                return parsed_dt.replace(tzinfo=datetime.timezone.utc).isoformat()
            except ValueError: continue
    log.debug(f"Could not parse PDF date '{date_str}'. Storing raw.")
    return date_str


# --- Base Class ---
class FileSystemItem(abc.ABC):
    """Abstract base class for filesystem items."""

    def __init__(self, path: Path, r_schijf_pad: Path, relative_path_override: Optional[str] = None):
        """Initializes FileSystemItem, ensuring paths are resolved."""
        self.path = path.resolve() # Work with resolved absolute paths
        self.r_schijf_pad = r_schijf_pad.resolve()
        self.name = self.path.name
        self._relative_path_override = relative_path_override
        self._metadata_cache: Optional[Dict[str, Any]] = None
        if not self.path.exists(): log.warning(f"Path DNE on init: {self.path}")

    @property
    @abc.abstractmethod
    def item_type(self) -> str: pass # Returns type identifier string

    @property
    def relative_path(self) -> str:
         """Calculates or returns the relative path string."""
         if self._relative_path_override is not None: return self._relative_path_override
         try: return str(self.path.relative_to(self.r_schijf_pad))
         except ValueError: return str(self.path) # Fallback if not relative
         except Exception as e: log.error(f"Error calc rel path {self.path}: {e}"); return str(self.path)

    @abc.abstractmethod
    def _extract_specific_metadata(self) -> Dict[str, Any]:
        """Subclass implementation for format-specific metadata."""
        pass

    def get_metadata(self) -> Optional[Dict[str, Any]]:
        """Gets metadata. Returns None only if path DNE or unreadable. Errors stored internally."""
        if self._metadata_cache is not None: return self._metadata_cache

        log.debug(f"Extracting ({self.item_type}): {self.path}")
        metadata: Dict[str, Any] = {
            "bestandspad": str(self.path), "bestandsnaam": self.name,
            "bestandstype": self.item_type, "relative_path": self.relative_path,
            "metadata": {}, "_extraction_timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        }
        specific_meta: Dict[str, Any] = {}
        error_msg: Optional[str] = None

        try:
            if not self.path.exists(): raise FileNotFoundError(f"Path vanished: {self.path}")
            if not os.access(self.path, os.R_OK): raise PermissionError(f"No read permission: {self.path}")

            # Common Stats
            try:
                 stat = self.path.stat()
                 specific_meta["file_size"] = stat.st_size
                 specific_meta["last_modified"] = _safe_isoformat(datetime.datetime.fromtimestamp(stat.st_mtime, datetime.timezone.utc))
            except Exception as stat_e: log.warning(f"Stat failed for {self.path}: {stat_e}")

            # Specific Extraction
            specific_meta.update(self._extract_specific_metadata())
            if err := specific_meta.pop("_extraction_error", None): error_msg = str(err)

            # Hash (Files only)
            if self.path.is_file():
                 file_hash = calculate_file_hash(self.path)
                 if file_hash: specific_meta["file_hash"] = file_hash
                 else: log.warning(f"Hash failed for: {self.path}"); specific_meta["file_hash"] = None

        except FileNotFoundError as e: log.error(str(e)); return None # Critical
        except PermissionError as e: log.error(str(e)); error_msg = "Permission denied"
        except Exception as e: log.exception(f"Unexpected error getting metadata for {self.path}: {e}"); error_msg = f"Unexpected error: {e}"

        # Assemble & Cache
        metadata["metadata"] = specific_meta
        if error_msg: metadata["metadata"]["_extraction_error"] = error_msg
        self._metadata_cache = metadata
        return self._metadata_cache

# --- Concrete Subclasses --- (Simplified v5)

class DirectoryItem(FileSystemItem):
    @property
    def item_type(self) -> str: return "directory"
    def _extract_specific_metadata(self) -> Dict[str, Any]: return {}

class PdfFile(FileSystemItem):
    @property
    def item_type(self) -> str: return "PDF"
    def _extract_specific_metadata(self) -> Dict[str, Any]:
        meta: Dict[str, Any] = {}
        try:
            with self.path.open("rb") as f:
                reader = PyPDF2.PdfReader(f, strict=False)
                if doc_info := reader.metadata:
                    meta.update({
                        "titel": str(doc_info.get('/Title','')).strip() or None,
                        "auteur": str(doc_info.get('/Author','')).strip() or None,
                        "onderwerp": str(doc_info.get('/Subject','')).strip() or None,
                        "producent": str(doc_info.get('/Producer','')).strip() or None,
                        "creatiedatum": _parse_pdf_date_simple(doc_info.get("/CreationDate")),
                        "wijzigingsdatum": _parse_pdf_date_simple(doc_info.get("/ModDate")),
                    })
                    # XMP
                    try:
                        if md_ref := doc_info.get('/Metadata'):
                            # Use get_object for potentially indirect objects
                            xmp_obj = reader.get_object(md_ref) if hasattr(reader, 'get_object') else None
                            if xmp_obj and hasattr(xmp_obj, 'get_data'):
                                data = xmp_obj.get_data()
                                try: meta["xmp"] = data.decode('utf-8')
                                except UnicodeDecodeError:
                                    try: enc = chardet.detect(data)['encoding'] or 'utf-8'; meta["xmp"] = data.decode(enc, errors='replace')
                                    except Exception: meta["xmp"] = None
                    except Exception as x_e: log.warning(f"XMP error {self.path}: {x_e}")
                try: meta["aantal_paginas"] = len(reader.pages) if not reader.is_encrypted else None
                except Exception: meta["aantal_paginas"] = None
        except PyPDF2.errors.PdfReadError as e: meta["_extraction_error"] = f"PDF Read Error: {e}"
        except Exception as e: meta["_extraction_error"] = f"PDF Error: {e}"; log.exception(f"PDF Error {self.path}")
        return meta

class ExcelFile(FileSystemItem):
    @property
    def item_type(self) -> str: return "Excel"
    def _extract_specific_metadata(self) -> Dict[str, Any]:
        meta: Dict[str, Any] = {}
        wb = None
        try:
            wb = load_workbook(filename=self.path, read_only=True, data_only=True, keep_links=False)
            meta["werkbladen"] = wb.sheetnames or []
            meta["aantal_werkbladen"] = len(meta["werkbladen"])
            if props := wb.properties:
                meta.update({
                    "titel": props.title, "auteur": props.creator, "onderwerp": props.subject,
                    "tags": props.keywords, "categorie": props.category,
                    "laatst_gewijzigd_door": props.lastModifiedBy,
                    "creatiedatum": _safe_isoformat(props.created),
                    "wijzigingsdatum": _safe_isoformat(props.modified)
                })
        except zipfile.BadZipFile: meta["_extraction_error"] = "Invalid XLSX format"
        except Exception as e: meta["_extraction_error"] = f"Excel Error: {e}"; log.exception(f"Excel Error {self.path}")
        finally:
            if wb: wb.close()
        return meta

class ShapefileFile(FileSystemItem):
    @property
    def item_type(self) -> str: return "Shapefile"
    def _extract_specific_metadata(self) -> Dict[str, Any]:
        if not GEOPANDAS_AVAILABLE: return {"_extraction_error": "Geopandas not installed"}
        meta: Dict[str, Any] = {}
        try:
            gdf = geopandas.read_file(self.path)
            meta.update({
                "crs": str(gdf.crs) if gdf.crs else None,
                "bounds": gdf.total_bounds.tolist(),
                "aantal_features": len(gdf),
                "kolomnamen": gdf.columns.tolist(),
                "geometrie_types": gdf.geometry.type.unique().tolist() if gdf.geometry is not None and not gdf.geometry.empty else []
            })
        except ImportError: meta["_extraction_error"] = "Geopandas import failed" # Defensive
        except Exception as e: meta["_extraction_error"] = f"Shapefile Error: {e}"; log.exception(f"Shapefile Error {self.path}")
        return meta

class ZipFileItem(FileSystemItem):
    @property
    def item_type(self) -> str: return "ZIP"
    def _extract_specific_metadata(self) -> Dict[str, Any]:
        meta: Dict[str, Any] = {}
        try:
             with zipfile.ZipFile(self.path, 'r') as zf:
                 if zf.comment:
                      try: meta["comment"] = zf.comment.decode('utf-8', errors='replace')
                      except Exception as e: log.warning(f"Zip comment decode error {self.path}: {e}")
        except zipfile.BadZipFile: meta["_extraction_error"] = "Invalid ZIP format"
        except Exception as e: meta["_extraction_error"] = f"ZIP Meta Error: {e}"; log.exception(f"ZIP Meta Error {self.path}")
        return meta

# --- Handler Mapping ---
SUPPORTED_FILE_EXTENSIONS = {
    ".pdf": PdfFile, ".xlsx": ExcelFile, ".xls": ExcelFile, ".shp": ShapefileFile,
}

# --- Metadata Scanner (Sequential Scan, Optional Parallelism ONLY within Zips) ---
class MetadataScanner:
    """Scans directories sequentially, extracts metadata, handles ZIPs recursively."""

    def __init__(self, root_scan_path: Union[str, Path], temp_dir: Union[str, Path] = DEFAULT_TEMP_DIR, max_workers_zip: Optional[int] = DEFAULT_MAX_WORKERS_ZIP):
        """Initializes MetadataScanner."""
        self.r_schijf_pad = Path(root_scan_path).resolve()
        self.temp_dir_base = Path(temp_dir).resolve()
        # Workers used ONLY for processing files *within* a zip file
        self.max_workers_zip = max_workers_zip if max_workers_zip and max_workers_zip > 0 else None
        # Public attribute to hold the path of the temp dir created for the current scan run
        # This path MUST be cleaned up by the caller (e.g., main.py) AFTER processing is complete.
        self.run_temp_dir: Optional[Path] = None

        if not self.r_schijf_pad.is_dir(): raise ValueError(f"Invalid root path: {self.r_schijf_pad}")
        log.info(f"Scanner Initialized. Root: {self.r_schijf_pad}, Temp Base: {self.temp_dir_base}, Zip Workers: {self.max_workers_zip or 'Sequential'}")

    def _create_item(self, path: Path, relative_override: Optional[str] = None) -> Optional[FileSystemItem]:
        """Factory method to create FileSystemItem."""
        try:
            p_resolved = path.resolve()
            if not p_resolved.exists(): return None
        except OSError as e: log.error(f"OS error checking path '{path}': {e}"); return None

        cls: Optional[Type[FileSystemItem]] = None
        if p_resolved.is_dir(): cls = DirectoryItem
        elif p_resolved.is_file():
            ext = p_resolved.suffix.lower()
            if ext == ".zip": cls = ZipFileItem
            else: cls = SUPPORTED_FILE_EXTENSIONS.get(ext)

        if cls:
             try: return cls(p_resolved, self.r_schijf_pad, relative_override)
             except Exception as e: log.exception(f"Init {cls.__name__} failed for {p_resolved}: {e}")
        elif p_resolved.is_file(): log.debug(f"Unsupported file: {p_resolved}")
        return None

    def _extract_zip_recursive(self, zip_path: Path, extract_dir: Path, processed: set) -> bool:
        """Extracts ZIP recursively. Returns True on success, False on critical failure."""
        abs_zip_path = zip_path.resolve()
        if abs_zip_path in processed: return True
        processed.add(abs_zip_path)
        log.debug(f"Extracting {zip_path} -> {extract_dir}")
        try:
            extract_dir.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(zip_path, 'r') as zf:
                 for member in zf.infolist():
                      try:
                           target = (extract_dir / member.filename).resolve()
                           if not str(target).startswith(str(extract_dir.resolve())):
                                log.warning(f"Skipping unsafe member: {member.filename} in {zip_path}")
                                continue
                           # Create parent directories for the member if they don't exist
                           target.parent.mkdir(parents=True, exist_ok=True)
                           # Now extract
                           zf.extract(member, path=extract_dir)
                      except Exception as e: log.error(f"Err extracting member {member.filename}: {e}") # Log & Continue

            all_nested_ok = True
            for item in extract_dir.rglob("*.zip"):
                 if item.is_file():
                      nested_dir = extract_dir / f"{item.stem}_nested_{abs_zip_path.name}"
                      if item.resolve() not in processed: # Check before recursive call
                         if not self._extract_zip_recursive(item, nested_dir, processed.copy()): all_nested_ok = False
                      else: log.warning(f"Skipping nested extraction of already processed: {item}")
                      try: item.unlink(missing_ok=True)
                      except OSError as e: log.warning(f"Failed nested zip cleanup {item}: {e}")
            return all_nested_ok
        except (zipfile.BadZipFile, FileNotFoundError, PermissionError, OSError) as e: log.error(f"Extraction failed for {zip_path}: {e}"); return False
        except Exception as e: log.exception(f"Unexpected error extracting {zip_path}: {e}"); return False


    def _process_extracted_item(self, item_path: Path, base_extract_dir: Path,
                                zip_item: ZipFileItem) -> Optional[Dict[str, Any]]:
        """Processes a single file found inside an extracted zip."""
        try:
            # Check if item still exists (might have failed extraction partially)
            if not item_path.exists():
                 log.warning(f"Extracted item vanished before processing: {item_path}")
                 return None
            rel_in_zip = str(item_path.relative_to(base_extract_dir))
            combined_rel = f"{zip_item.relative_path}/{rel_in_zip}"
            fs_item = self._create_item(item_path, relative_override=combined_rel)
            if not fs_item: return None

            metadata = fs_item.get_metadata() # Get metadata from the temp path
            if metadata:
                # Add context, but keep original bestandspad pointing to temp file
                metadata["metadata"]["_source_zip_file"] = str(zip_item.path)
                metadata["metadata"]["_temp_extraction_path"] = metadata["bestandspad"] # Store temp path explicitly
                metadata["metadata"]["potential_organization_id"] = zip_item.path.stem.lower().replace(" ", "-").replace("_", "-")
                # Override bestandspad to reflect the original logical path? This is complex.
                # For now, keep bestandspad as the temp path, upload uses this path.
                return metadata
            else: log.warning(f"Metadata failed for extracted item: {item_path}")
        except ValueError as e: log.error(f"Path error processing extracted {item_path}: {e}")
        except Exception as e: log.exception(f"Error processing extracted {item_path}: {e}")
        return None


    def _process_zip_contents(self, zip_item: ZipFileItem) -> List[Dict[str, Any]]:
        """Extracts zip, processes contents (optionally parallel), returns metadata list."""
        if not self.run_temp_dir: log.error("Cannot process ZIP: main run temporary directory not set."); return []

        abs_zip_path = zip_item.path.resolve()
        extract_dir = self.run_temp_dir / f"{zip_item.path.stem}_{abs_zip_path.name}_contents"

        log.info(f"Processing contents of zip: {zip_item.path}")
        all_zip_meta = []
        try:
            if not self._extract_zip_recursive(zip_item.path, extract_dir, set()):
                 log.error(f"Aborting content processing due to extraction failure: {zip_item.path}"); return []

            items_to_process: List[Path] = list(extract_dir.rglob("*")) # Get all items recursively
            files_in_zip = [p for p in items_to_process if p.is_file() and not p.name.lower().endswith('.zip')] # Filter out nested zips already handled

            if files_in_zip:
                 log.info(f"Found {len(files_in_zip)} files in {zip_item.path}. Extracting metadata...")
                 process_func: Callable[[Path], Optional[Dict[str, Any]]] = \
                     lambda p: self._process_extracted_item(p, extract_dir, zip_item)

                 if self.max_workers_zip and len(files_in_zip) > 1:
                      log.debug(f"Processing zip contents in parallel (workers={self.max_workers_zip})")
                      with ThreadPoolExecutor(max_workers=self.max_workers_zip) as executor:
                           futures = {executor.submit(process_func, file_path): file_path for file_path in files_in_zip}
                           for future in as_completed(futures):
                                source_path = futures[future]
                                try:
                                     result = future.result()
                                     if result: all_zip_meta.append(result)
                                except Exception as e: log.exception(f"Error processing {source_path} from zip {zip_item.path} in parallel: {e}")
                 else: # Sequential processing within zip
                      log.debug("Processing zip contents sequentially.")
                      for file_path in files_in_zip:
                           result = process_func(file_path)
                           if result: all_zip_meta.append(result)
        except Exception as e:
            log.exception(f"Failed to process zip {zip_item.path} contents: {e}")
        finally:
            # Clean up THIS zip's specific temp dir immediately
            if extract_dir.exists(): shutil.rmtree(extract_dir, ignore_errors=True); log.debug(f"Cleaned zip temp: {extract_dir}")
        return all_zip_meta


    def scan(self) -> List[Dict[str, Any]]:
        """Performs a sequential scan, extracts metadata, handles ZIPs recursively."""
        all_metadata: List[Dict[str, Any]] = []
        processed_paths: set = set()

        log.info(f"Starting sequential scan: {self.r_schijf_pad}")
        ts = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%d_%H%M%S")
        # Set the run_temp_dir for this scan. It must be cleaned by the caller (main.py)
        self.run_temp_dir = self.temp_dir_base / f"scan_{ts}"

        try:
            self.run_temp_dir.mkdir(parents=True, exist_ok=True)
            log.info(f"Using run temp dir (MUST BE CLEANED LATER): {self.run_temp_dir}")

            for root, dirs, files in os.walk(self.r_schijf_pad, topdown=True, onerror=log.error):
                 root_path = Path(root)
                 # Combine dirs and files, process sequentially
                 for name in dirs + files:
                      item_path = root_path / name
                      try:
                          abs_path = item_path.resolve()
                          if abs_path in processed_paths: continue
                          processed_paths.add(abs_path)

                          item = self._create_item(item_path)
                          if not item: continue

                          if isinstance(item, ZipFileItem):
                               zip_meta = item.get_metadata() # Metadata for zip file itself
                               if zip_meta: all_metadata.append(zip_meta)
                               content_meta = self._process_zip_sequentially(item) # Extract and process contents
                               all_metadata.extend(content_meta)
                          else: # Regular Directory or File
                               item_meta = item.get_metadata()
                               if item_meta: all_metadata.append(item_meta)

                      except OSError as e: log.error(f"OS error processing {item_path}: {e}")
                      except Exception as e: log.exception(f"Unexpected error on {item_path}: {e}")

        except Exception as e:
            log.exception(f"Critical error during scan execution: {e}")
            # Do NOT cleanup temp dir here if scan fails, leave it for inspection? Or try? Let's try.
            self._try_cleanup_run_temp_dir()

        # --- CRITICAL: REMOVED automatic cleanup from scan() ---
        # The caller (main.py) is now responsible for cleaning self.run_temp_dir AFTER uploads

        log.info(f"Sequential scan complete. Found {len(all_metadata)} metadata entries.")
        return all_metadata

    def _try_cleanup_run_temp_dir(self):
        """Internal helper to attempt cleanup, called only on major scan failure."""
        if self.run_temp_dir and self.run_temp_dir.exists():
            log.warning(f"Attempting emergency cleanup of run temp dir due to scan error: {self.run_temp_dir}")
            shutil.rmtree(self.run_temp_dir, ignore_errors=True)
        self.run_temp_dir = None # Reset path


# --- Validation & Transformation Classes (Assumed OK from previous versions) ---

class MetadataValidator:
    """Validates extracted metadata dictionaries."""
    def validate(self, metadata: Dict[str, Any]) -> Dict[str, str]:
        results: Dict[str, str] = {}
        item_type = metadata.get("bestandstype"); spec_meta = metadata.get("metadata", {}); path = metadata.get('bestandspad','?')
        if err := spec_meta.get("_extraction_error"): results["_extraction_error"] = f"Extraction error: {err}"
        validators = {"PDF": self._v_pdf, "Excel": self._v_excel, "Shapefile": self._v_shp}
        if func := validators.get(item_type): results.update(func(spec_meta))
        if results: log.debug(f"Validation issues for {path}: {results}")
        return results
    def _v_pdf(self, m: Dict[str, Any]) -> Dict[str, str]:
        e={}; p=m.get("aantal_paginas"); t=m.get("titel")
        if not t: e["titel"]="PDF Title missing"
        if p is None: e["aantal_paginas"]="Page count N/A"
        elif isinstance(p,int) and p>1000: e["aantal_paginas"]=f"Too many pages ({p}>1000)"
        return e
    def _v_excel(self, m: Dict[str, Any]) -> Dict[str, str]:
        e={}; s=m.get("aantal_werkbladen"); t=m.get("titel")
        if not t: e["titel"]="Excel Title missing"
        if s is None: e["aantal_werkbladen"]="Sheet count N/A"
        elif isinstance(s,int) and s>50: e["aantal_werkbladen"]=f"Too many sheets ({s}>50)"
        return e
    def _v_shp(self, m: Dict[str, Any]) -> Dict[str, str]:
        e = {}
        if not m.get("crs"): e["crs"]="CRS missing"
        if not m.get("geometrie_types"): e["geometrie_types"]="Geom types missing"
        return e

class CkanTransformer:
    """Transforms extracted metadata for CKAN."""
    def __init__(self, ckan_mapping: Dict[str, str]): self.map = ckan_mapping
    def _slugify(self, t: str, ml: int = 100) -> str:
        if not t: return "unnamed"; s=str(t).lower().strip(); s=re.sub(r'[\s\._]+', '-', s); s=re.sub(r'[^\w-]', '', s); s=re.sub(r'-{2,}', '-', s); s=s.strip('-')[:ml]; return s or "unnamed"
    def transform(self, metadata: Dict[str, Any]) -> Dict[str, Any]:
        out: Dict[str, Any] = {}; spec = metadata.get("metadata", {}); p = metadata.get('bestandspad','?')
        for src, target in self.map.items():
            v = spec.get(src, metadata.get(src))
            if v is None: continue
            try: # Apply transformations
                if target == "tags" and isinstance(v, str): out[target] = [t.strip() for t in v.split(',') if t.strip()]
                elif target == "format" and isinstance(v, str): out[target] = v.upper()
                elif target == "spatial_bbox" and isinstance(v, list) and len(v) == 4: coords = [float(c) for c in v]; out["spatial"] = json.dumps({"type": "Polygon","coordinates": [[[coords[0], coords[1]], [coords[2], coords[1]], [coords[2], coords[3]], [coords[0], coords[3]], [coords[0], coords[1]]]]})
                else: out[target] = v # Direct map
            except Exception as e: log.warning(f"Transform error {src}->{target} for {p}: {e}"); out[f"_err_{target}"] = str(e)
        # Defaults
        if not out.get('title'): out['title'] = metadata.get('bestandsnaam', 'Untitled')
        if not out.get('name'): out['name'] = self._slugify(out['title'])
        log.debug(f"Transformed metadata for {p}")
        return out

# --- Example Usage --- (For testing this file directly)
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)-8s %(name)s: %(message)s')
    TEST_DIR = Path("./_test_scan_data_v5") # Use unique name for testing
    try:
        log.info(f"Setting up test dir {TEST_DIR}")
        TEST_DIR.mkdir(exist_ok=True)
        (TEST_DIR / "doc1.pdf").write_text("pdf content v5", encoding="utf-8")
        # Create a minimal valid xlsx for testing properties if possible
        try:
             wb_test = load_workbook() # Create empty
             wb_test.properties = type('obj', (object,), {'creator': 'Test Script', 'title': 'Test Sheet'})() # Dummy props
             wb_test.save(TEST_DIR / "sheet1.xlsx")
             wb_test.close()
        except Exception as xlsx_e: log.error(f"Could not create test xlsx: {xlsx_e}")

        zip_path = TEST_DIR / "data_v5.zip"
        with zipfile.ZipFile(zip_path, 'w') as zf:
             zf.writestr("inside_v5.txt", "zip content v5")
             zf.writestr("nested/readme.md", "# Readme v5")
        log.info("Test files created.")

        # Minimal mapping for example
        EXAMPLE_CKAN_MAPPING = {"bestandstype": "format", "titel": "title", "name": "name", "file_hash": "hash"}

        scanner = MetadataScanner(TEST_DIR) # Use default temp dir
        validator = MetadataValidator()
        transformer = CkanTransformer(EXAMPLE_CKAN_MAPPING)

        log.info("--- Starting Scan ---")
        all_meta = scanner.scan()
        log.info(f"--- Scan Finished ({len(all_meta)} items) ---")

        results = []
        for meta in all_meta:
            print("-" * 20)
            path = meta.get('relative_path')
            errors = validator.validate(meta)
            transformed = transformer.transform(meta)
            print(f"PATH: {path}")
            print(f"VALIDATION: {'OK' if not errors else errors}")
            print(f"TRANSFORMED: {json.dumps(transformed, indent=2)}")

    except Exception as main_e:
        log.exception(f"Error during example execution: {main_e}")
    finally:
        # Cleanup test dir
        if TEST_DIR.exists(): log.info(f"Cleaning up {TEST_DIR}"); shutil.rmtree(TEST_DIR, ignore_errors=True)
        # Also attempt cleanup of the scanner's run temp dir if path is known
        if 'scanner' in locals() and scanner.run_temp_dir and scanner.run_temp_dir.exists():
             log.info(f"Cleaning up scanner temp dir: {scanner.run_temp_dir}")
             shutil.rmtree(scanner.run_temp_dir, ignore_errors=True)