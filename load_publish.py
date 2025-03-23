import os
import logging
from typing import List, Dict, Any, Optional
import PyPDF2
from openpyxl import load_workbook
import geopandas
import chardet
from pathlib import Path
import utils
import zipfile  # Import zipfile


log = logging.getLogger(__name__)

def scan_en_extraheer(r_schijf_pad: str) -> List[Dict[str, Any]]:
    """Scans the R-drive, extracts files (including from ZIPs), and extracts metadata."""
    metadata_lijst: List[Dict[str, Any]] = []
    temp_dir = Path("./temp_extraction")
    if not temp_dir.exists():
        temp_dir.mkdir()

    try:
        for root, dirs, files in os.walk(r_schijf_pad):
            # --- Process Directories (for modification times) ---
            for dir_name in dirs:
                dir_path = Path(root) / dir_name
                metadata = {
                    "bestandspad": str(dir_path),
                    "bestandsnaam": dir_name,
                    "bestandstype": "directory",
                    "metadata": {
                        "last_modified": os.path.getmtime(dir_path),
                    },
                    "relative_path": str(dir_path.relative_to(r_schijf_pad)),
                }
                metadata_lijst.append(metadata)

            # --- Process Files ---
            for file_name in files:
                file_path = Path(root) / file_name
                try:
                    if file_path.suffix.lower() == ".zip":
                        # Extract ZIP file and process its contents
                        _extract_zip(file_path, temp_dir)
                        #Now, walk through newly extracted files:
                        for extracted_root, _, extracted_files in os.walk(temp_dir):
                            for extracted_file in extracted_files:
                                 extracted_path = Path(extracted_root) / extracted_file
                                 metadata = _extract_metadata(extracted_path)
                                 if metadata:
                                    metadata["relative_path"] = str(
                                        Path(root) / file_name / extracted_path.relative_to(temp_dir) #Correct relative path.
                                    )
                                    metadata["metadata"]["file_hash"] = utils.calculate_file_hash(str(extracted_path))
                                    metadata_lijst.append(metadata)
                                 extracted_path.unlink() #Delete extracted file
                        #Clean up the temporary directory and its contents after processing the zip.
                        for temp_root, temp_dirs, temp_files in os.walk(temp_dir, topdown=False):
                            for temp_file in temp_files:
                                (Path(temp_root) / temp_file).unlink()
                            for temp_dir_name in temp_dirs:
                                (Path(temp_root) / temp_dir_name).rmdir()

                    else:  # Regular file
                        metadata = _extract_metadata(file_path)
                        if metadata:
                            metadata["relative_path"] = str(file_path.relative_to(r_schijf_pad))
                            metadata["metadata"]["file_hash"] = utils.calculate_file_hash(str(file_path))
                            metadata_lijst.append(metadata)
                except Exception as e:
                    log.error(f"Fout bij verwerken van {file_path}: {e}", exc_info=True)
    finally:
        # Clean up the temporary directory ONLY IF it's empty
        if temp_dir.exists() and not any(temp_dir.iterdir()):
            temp_dir.rmdir()

    return metadata_lijst

def _extract_zip(zip_path: Path, extract_to: Path):
    """Recursively extracts ZIP files, including nested ZIPs."""
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)

        # Check for nested ZIP files
        for item in extract_to.glob("**/*.zip"):  # Find any .zip files recursively
            if item.is_file():
                nested_temp_dir = extract_to / item.stem  # Create a subdir for the nested zip
                if not nested_temp_dir.exists():
                    nested_temp_dir.mkdir()
                _extract_zip(item, nested_temp_dir)  # Recursive call
                item.unlink() # Delete after extraction

    except zipfile.BadZipFile:
        log.error(f"Corrupt ZIP file: {zip_path}")
    except Exception as e:
        log.error(f"Error extracting ZIP file {zip_path}: {e}", exc_info=True)

def _extract_metadata(bestandspad: Path) -> Optional[Dict[str, Any]]:
    """Extraheert metadata op basis van bestandstype (interne hulpfunctie)."""
    metadata = {
        "bestandspad": str(bestandspad),
        "bestandsnaam": bestandspad.name,
        "bestandstype": None,
        "metadata": {}
    }
    bestands_extensie = bestandspad.suffix.lower()

    if bestands_extensie == ".pdf":
        metadata["bestandstype"] = "PDF"
        metadata["metadata"] = _extract_pdf_metadata(bestandspad)
    elif bestands_extensie in (".xlsx", ".xls"):
        metadata["bestandstype"] = "Excel"
        metadata["metadata"] = _extract_excel_metadata(bestandspad)
    elif bestands_extensie == ".shp":
        metadata["bestandstype"] = "Shapefile"
        metadata["metadata"] = _extract_shapefile_metadata(bestandspad)
    elif bestands_extensie == ".zip": #Skip zips, because we handle them separately
        return None
    else:
        log.warning(f"Onbekend bestandstype: {bestandspad}")
        return None

    return metadata

def _extract_pdf_metadata(bestandspad: Path) -> Dict[str, Any]:
    """Extraheert metadata uit een PDF."""
    metadata: Dict[str, Any] = {}
    try:
        with open(bestandspad, "rb") as f:
            pdf_reader = PyPDF2.PdfReader(f)
            doc_info = pdf_reader.metadata
            metadata["titel"] = doc_info.get("/Title")
            metadata["auteur"] = doc_info.get("/Author")
            metadata["onderwerp"] = doc_info.get("/Subject")
            metadata["creatiedatum"] = doc_info.get("/CreationDate")
            metadata["wijzigingsdatum"] = doc_info.get("/ModDate")
            metadata["producent"] = doc_info.get("/Producer")
            metadata["aantal_paginas"] = len(pdf_reader.pages)

            xmp_metadata = doc_info.get("/Metadata")
            if xmp_metadata:
                try:
                    encoding = chardet.detect(xmp_metadata.get_data())['encoding'] or 'utf-8'
                    metadata["xmp"] = xmp_metadata.get_data().decode(encoding, errors='replace')
                except Exception as e:
                    log.warning(f"Fout bij decoderen XMP metadata in {bestandspad}: {e}")
                    metadata["xmp"] = None

    except (PyPDF2.errors.PdfReadError, Exception) as e:
        log.error(f"Fout bij extractie PDF metadata van {bestandspad}: {e}", exc_info=True)
        metadata["fout"] = "Corrupte PDF of andere fout"
    return metadata

def _extract_excel_metadata(bestandspad: Path) -> Dict[str, Any]:
    """Extraheert metadata uit een Excel."""
    metadata: Dict[str, Any] = {}
    try:
        workbook = load_workbook(filename=bestandspad, read_only=True)
        metadata["werkbladen"] = workbook.sheetnames
        metadata["aantal_werkbladen"] = len(workbook.sheetnames)
        props = workbook.properties
        metadata["titel"] = props.title
        metadata["auteur"] = props.creator
        metadata["laatst_gewijzigd_door"] = props.lastModifiedBy
        metadata["creatiedatum"] = props.created
        metadata["wijzigingsdatum"] = props.modified
        metadata["onderwerp"] = props.subject
        metadata["categorie"] = props.category
        metadata["tags"] = props.keywords
    except Exception as e:
        log.error(f"Fout bij extractie Excel metadata van {bestandspad}: {e}", exc_info=True)
        metadata["fout"] = "Fout bij lezen Excel"
    return metadata

def _extract_shapefile_metadata(bestandspad: Path) -> Dict[str, Any]:
    """Extraheert metadata uit een Shapefile."""
    metadata: Dict[str, Any] = {}
    try:
        gdf = geopandas.read_file(bestandspad)
        metadata["crs"] = str(gdf.crs)
        metadata["bounds"] = gdf.total_bounds.tolist()
        metadata["aantal_features"] = len(gdf)
        metadata["kolomnamen"] = gdf.columns.tolist()
        metadata["geometrie_types"] = gdf.geometry.type.unique().tolist()
    except Exception as e:
        log.error(f"Fout bij extractie Shapefile metadata van {bestandspad}: {e}", exc_info=True)
        metadata["fout"] = "Fout bij lezen Shapefile"
    return metadata


def valideer_metadata(metadata: Dict[str, Any]) -> Dict[str, str]:
    """Valideert metadata."""
    validatie_resultaten: Dict[str, str] = {}
    bestandstype = metadata.get("bestandstype")

    if bestandstype == "PDF":
        if not utils.extract_value(metadata, "titel", "metadata"):
            validatie_resultaten["titel"] = "Ontbreekt"
        if (aantal_paginas := utils.extract_value(metadata, "aantal_paginas", "metadata")) and aantal_paginas > 1000:
             validatie_resultaten["aantal_paginas"] = "Te veel pagina's"

    elif bestandstype == "Excel":
        if (aantal_werkbladen:= utils.extract_value(metadata, "aantal_werkbladen", "metadata")) and aantal_werkbladen > 25:
            validatie_resultaten["aantal_werkbladen"] = "Te veel werkbladen"
        if not utils.extract_value(metadata,"titel", "metadata"):
            validatie_resultaten["titel"] = "Ontbreekt"

    elif bestandstype == "Shapefile":
        if not utils.extract_value(metadata, "crs", "metadata"):
            validatie_resultaten["crs"] = "Ontbreekt"
        if not utils.extract_value(metadata, "geometrie_types", "metadata"):
            validatie_resultaten["geometrie_types"] = "Geen geometrie types"

    return validatie_resultaten


def transformeer_naar_ckan(metadata: Dict[str, Any], ckan_mapping: Dict[str, str]) -> Dict[str, Any]:
    """Transformeert metadata naar CKAN formaat."""
    ckan_metadata: Dict[str, Any] = {}

    for r_veld, ckan_veld in ckan_mapping.items():
        waarde = utils.extract_value(metadata, r_veld, "metadata") or metadata.get(r_veld)
        if waarde is not None:
            if ckan_veld == "keywords" and isinstance(waarde, str):
                waarde = waarde.split(",")
            elif ckan_veld == "format":
                waarde = waarde.upper()
            elif ckan_veld == "spatial_bbox" and isinstance(waarde, list):
                ckan_metadata["spatial"] = {
                    "type": "Polygon",
                    "coordinates": [[[waarde[0], waarde[1]], [waarde[2], waarde[1]],
                                     [waarde[2], waarde[3]], [waarde[0], waarde[3]],
                                     [waarde[0], waarde[1]]]]
                }
            else:
                ckan_metadata[ckan_veld] = waarde
    return ckan_metadata